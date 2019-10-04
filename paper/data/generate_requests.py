from openpyxl import Workbook
from datetime import datetime, timedelta
import itertools

num_clin = 10
req_per_clin = 10

def main(num_clinicians, req_per_clin):
    start = datetime(2019, 1, 1)
    next_date = datetime(2019, 1, 1)
    blocks = []
    clin_requests = {i: list() for i in range(1, num_clin + 1)}

    while (next_date - start).days < 360:
        blocks.append(next_date)
        next_date = next_date + timedelta(weeks=2)

    wb = Workbook()
    block_iterator = itertools.cycle(blocks)
    for clin in range(num_clinicians):
        ws = wb.create_sheet(str(clin + 1))
        for row in range(1, req_per_clin + 1):
            val = next(block_iterator).strftime('%d-%b-%Y')
            ws.cell(column=1, row=row, value=val)
            ws.cell(column=2, row=row, value=val)

    wb.remove(wb.active)
    wb.save('{}clin_{}requests.xlsx'.format(num_clinicians, req_per_clin))

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('num_clinicians', type=int)
    parser.add_argument('req_per_clinician', type=int)

    args = parser.parse_args()
    main(args.num_clinicians, args.req_per_clinician)