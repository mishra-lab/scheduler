import calendar
import itertools
import operator
from collections import namedtuple
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, colors
from openpyxl.utils import get_column_letter

from constants import WEEK_HOURS, WEEKEND_HOURS


class ExcelHelper:
    @staticmethod
    def saveYearlySchedule(filename, table):
        wb = Workbook()
        sheet = wb.active

        cols = table.columnCount()
        rows = table.rowCount()

        # write header
        for i in range(cols):
            header = table.horizontalHeaderItem(i)
            sheet.cell(row=1, column=i + 1, value=header.text())

        # write data
        sheetRow = 2
        for i in range(rows):
            for j in range(cols):
                item = table.item(i, j)
                sheet.cell(row=sheetRow, column=j+1, value=item.text())
            sheetRow += 1

        wb.save(filename)

    @staticmethod
    def saveMonthlySchedule(filename, table, year, holidayMap):
        # for each weeknum, loop over days in that week, assuming we start
        # at monday and end on friday
        
        DayFields = namedtuple('DayFields', 
            ['fulldate', 'date', 'day', 'dayclin', 'nightclin'])

        cal = calendar.Calendar()
        numCols = table.columnCount()
        divisions = ExcelHelper.getHorizontalHeaders(table)[1:-1]
        weekNums = ExcelHelper.getColumn(table, 0)
        clinicianTuples = ExcelHelper.getColumns(table, [i for i in range(1, numCols)])

        def expand_dates(weekText, clinTuple):
            '''
            Given a week number and a set of clinicians that cover the corresponding \
            week + weekend, return a mapping between weekday and clinician(s) covering that day
            '''

            result = []
            weekNum = int(
                weekText[:-1]) if weekText[-1] == '*' else int(weekText)

            # take the last clinician to be the one on-call for the weekend
            weekClinicians = clinTuple[:-1]
            weekendClinician = tuple([clinTuple[-1]] * len(divisions))

            day = weekStart = datetime.strptime(
                # year / weekNum / Monday / 8AM
                '{0}/{1:02d}/1/08:00'.format(year, weekNum),
                '%G/%V/%u/%H:%M'
            )
            weekEnd = weekStart + timedelta(hours=WEEK_HOURS)

            # enumerate weekdays
            while day < weekEnd:
                result.append(DayFields(
                    '{:%Y-%m-%d}'.format(day), 
                    '{:%b-%d}'.format(day), 
                    '{:%a}'.format(day),
                    weekClinicians,

                    # handle friday separately, distinguish day/night clinicians
                    weekendClinician if day.date() == weekEnd.date() else weekClinicians
                ))
                day += timedelta(days=1)

            # enumerate weekend
            nextStart = weekEnd + timedelta(hours=WEEKEND_HOURS)
            while day < nextStart:
                result.append(DayFields(
                    '{:%Y-%m-%d}'.format(day), 
                    '{:%b-%d}'.format(day), 
                    '{:%a}'.format(day), 
                    weekendClinician,
                    weekendClinician
                ))
                day += timedelta(days=1)

            return result

        def group_months(tup):
            '''
            Given a DayFields tuple, figure out which month it belongs in.

            Note: months will always have complete weeks, so this will group together dates
            that may be outside of a certain month (e.g.: if Dec 31 is a Monday, it will be grouped together
            with January)
            '''
            date = tup.fulldate
            
            for i in range(1, 13):
                # search for date in month iterator
                it = cal.itermonthdates(year, i)
                month_dates = list(map(lambda x: '{:%Y-%m-%d}'.format(x), it))

                if date in month_dates:
                    return (year, i)

        # create clinician mapping for each day of the year based on week numbers
        date_maps = list(map(expand_dates, weekNums, clinicianTuples))

        # group day mappings according to month
        all_dates = list(itertools.chain.from_iterable(date_maps))

        # deal with holidays after we created all_dates
        for i in range(len(all_dates)):
            date = all_dates[i]
            if date.fulldate in holidayMap:
                weekNum = holidayMap[date.fulldate]
                clin = clinicianTuples[weekNum - 1][-1]

                # replace date with the correct DayTuple
                newdate = date._replace(dayclin=(clin, clin), nightclin=(clin, clin))
                all_dates[i] = newdate

        groupby_months = itertools.groupby(all_dates, group_months)

        month_dict = {k : list(v) for k, v in groupby_months}
        keys = sorted(month_dict.keys(), key=operator.itemgetter(0, 1))

        # write month_dict to excel, separate sheet for each month
        wb = Workbook()
        for key in keys:
            title = calendar.month_name[key[1]]
            ws = wb.create_sheet(title)

            # write header
            ws.cell(row=2, column=1, value='Date')
            ws.cell(row=2, column=2, value='Day')
            ws.cell(row=2, column=3, value='Fellow')

            col_idx = 4
            for div in divisions:
                ExcelHelper.horizontalCenter(
                    ws.cell(row=1, column=col_idx, value='{} Service'.format(div))
                )
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+2)
                ws.cell(row=2, column=col_idx, value='0800 - 1700')
                ws.cell(row=2, column=col_idx+1, value='1700 - 0800')
                ws.cell(row=2, column=col_idx+2, value='Backup')
                col_idx += 3

            # write month sheet
            month = month_dict[key]
            for i in range(len(month)):
                daytuple = month[i]

                ExcelHelper.addFullBlackBorder(
                    ws.cell(row=(i+3), column=1, value=daytuple.date)
                )
                ExcelHelper.addFullBlackBorder(
                    ws.cell(row=(i+3), column=2, value=daytuple.day)
                )
                ExcelHelper.horizontalCenter(
                    ExcelHelper.addFullBlackBorder(
                        ws.cell(row=(i+3), column=3, value=None)
                    )
                )

                col_idx = 4
                for j in range(len(divisions)):
                    dayClinician = daytuple.dayclin[j]
                    nightClinician = daytuple.nightclin[j]

                    # use next division's night clinician as a backup
                    next_div = (j + 1) % len(divisions)
                    backupClinician = daytuple.nightclin[next_div]

                    ExcelHelper.horizontalCenter(
                        ExcelHelper.addFullBlackBorder(
                            ws.cell(row=(i+3), column=col_idx, value=dayClinician)
                        )
                    )
                    ExcelHelper.horizontalCenter(
                        ExcelHelper.addFullBlackBorder(
                            ws.cell(row=(i+3), column=col_idx+1, value=nightClinician)
                        )
                    )
                    ExcelHelper.horizontalCenter(
                        ExcelHelper.addFullBlackBorder(
                            ws.cell(row=(i+3), column=col_idx+2, value=backupClinician)
                        )
                    )

                    col_idx += 3

            ExcelHelper.expandColumns(ws)
            
        # remove the automatically created first sheet
        wb.remove(wb.active)
        wb.save(filename)

    @staticmethod
    def getColumn(table, col_idx):
        rows = table.rowCount()
        column = []
        for i in range(rows):
            column.append(table.item(i, col_idx).text())

        return column

    @staticmethod
    def getColumns(table, col_idxs):
        rows = table.rowCount()
        result = []
        for i in range(rows):
            row = []
            for j in col_idxs:
                row.append(table.item(i, j).text())

            result.append(tuple(row))

        return result

    @staticmethod
    def getRow(table, row_idx):
        cols = table.columnCount()
        row = []
        for i in range(cols):
            row.append(table.item(row_idx, i).text())

        return row

    @staticmethod
    def getHorizontalHeaders(table):
        cols = table.columnCount()
        headers = []
        for i in range(cols):
            headers.append(table.horizontalHeaderItem(i).text())

        return headers

    @staticmethod
    def expandColumns(worksheet, expand_factor=1.25):
        """
        Expands every column width in given worksheet up to 
            max_length * expand_factor

        ref: https://groups.google.com/d/msg/openpyxl-users/rsy8W2epzVs/a0sjdghwCAAJ
        """

        for idx, col in enumerate(worksheet.columns, 1):
            lengths = [len(u"{}".format(cell.value)) for cell in col if cell.value != None]
            new_width = max(lengths) * expand_factor
            worksheet.column_dimensions[get_column_letter(idx)].width = new_width

        return worksheet

    @staticmethod
    def addFullBlackBorder(cell):
        side = Side(border_style='thin', color=colors.BLACK)
        cell.border = Border(left=side, right=side, top=side, bottom=side)
        return cell

    @staticmethod
    def horizontalCenter(cell):
        cell.alignment = Alignment(horizontal='center')
        return cell