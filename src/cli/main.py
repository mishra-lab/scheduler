import os
import time
from datetime import datetime, timedelta

from openpyxl import Workbook

import scheduler
from constants import *
from oauth2client import tools


def publish_sched(sched, api, year):
    """
    Publishes the block and weekend assignments as events to Google 
    calendar.
    """
    for division in sched.divisions.values():
        for block_num in range(len(division.assignments)):
            clinician = division.assignments[block_num]
            for j in range(BLOCK_SIZE, 0, -1):
                week_start = datetime.strptime(
                    '{0}/{1:02d}/1/08:00/'.format(
                        year,
                        BLOCK_SIZE * (block_num + 1) - (j - 1)),
                    '%G/%V/%u/%H:%M/')
                week_end = week_start + timedelta(hours=WEEK_HOURS)
                summary = '[scheduler] {} - DIV:{} on call'.format(
                    clinician.name, division.name)
                api.create_event(
                    week_start.isoformat(),
                    week_end.isoformat(),
                    [clinician.email],
                    summary
                )

    for clinician in sched.clinicians.values():
        for week_num in clinician.weekends_assigned:
            weekend_start = datetime.strptime(
                '{0}/{1:02d}/5/17:00/'.format(
                    year,
                    week_num),
                '%G/%V/%u/%H:%M/')
            weekend_end = weekend_start + timedelta(hours=WEEKEND_HOURS)
            summary = '[scheduler] {} - on call'.format(clinician.name)
            api.create_event(
                weekend_start.isoformat(),
                weekend_end.isoformat(),
                [clinician.email],
                summary
            )


def write_to_excel(scheduler, year):
    wb = Workbook()
    sheet = wb.active
    col_index = 0
    divisions = list(scheduler.divisions.values())

    for col_index in range(len(divisions)):
        division = divisions[col_index]
        sheet.cell(1, col_index + 1, division.name)
        for row_index in range(len(division.assignments)):
            clinician = division.assignments[row_index]
            sheet.cell(2 * (row_index + 1), col_index + 1, clinician.name)
            sheet.cell(2 * (row_index + 1) + 1, col_index + 1, clinician.name)

    weekend_assignments = [
        _ for clinician in scheduler.clinicians.values()
        for _ in clinician.get_weekend_vars(
            lambda x: x.get_value() == 1.0
        )
    ]
    weekend_assignments.sort(key=lambda x: x.week_num)
    col_index += 1
    sheet.cell(1, col_index + 1, "Weekends")
    for wa in weekend_assignments:
        if wa.week_num in scheduler.long_weekends:
            sheet.cell((wa.week_num + 1), col_index +
                       1, wa.clinician.name + "****")
        else:
            sheet.cell((wa.week_num + 1), col_index + 1, wa.clinician.name)

    wb.save('{}-schedule.xlsx'.format(year))


def generate_schedule(args):
    global NUM_BLOCKS
    if args.blocks: NUM_BLOCKS = args.blocks

    sched = scheduler.Scheduler(args.config, NUM_BLOCKS)
    requests_off = []
    long_weekends = set()

    if args.calendar:
        print("Retrieving {} calendar events from {}...".format(args.year, args.calendar))
        api = scheduler.API(args.calendar, args)
        start_date = datetime(args.year, 1, 1)
        end_date = start_date + timedelta(weeks=52)
        events = api.get_events(
            start_date.isoformat() + 'Z', end_date.isoformat() + 'Z')

        requests_off = list(filter(lambda x: '[request] ' in x['summary'], events))
        # populate long weekends
        lw_events = list(filter(lambda x: '[holiday] ' in x['summary'], events))
        for evt in lw_events:
            start = datetime.strptime(
                evt['start'].get('date'),
                '%Y-%m-%d'
            )

            # Fri statutory holidays are associated with their regular weeknum
            # Mon statutory holidays are associated with their weeknum - 1
            #   (i.e.: the previous weeknum)
            if start.isoweekday() == 1:
                long_weekends.add(start.isocalendar()[1] - 1)
            elif start.isoweekday() == 5:
                long_weekends.add(start.isocalendar()[1])

    print("Populating scheduler...")
    sched.set_timeoff(requests_off)
    sched.set_long_weekends(list(long_weekends))
    sched.build_lp()
    if sched.solve_lp():
        print("Found a feasible schedule!")
        sched.assign_schedule()
        write_to_excel(sched, args.year)

        if args.publish:
            print("Publishing {} schedule to {}...".format(args.year, args.calendar))
            publish_sched(sched, api, args.year)
    else:
        print("ERROR: Could not find a feasible schedule.")
        print("Try adjusting min/max values for clinicians.")


def clear_schedule(args):
    api = scheduler.API(args.calendar, args)
    start_date = datetime(args.year, 1, 1)
    end_date = start_date + timedelta(weeks=52)

    print("Clearing previously generated {} schedule from {}...".format(args.year, args.calendar))
    api.delete_events(start_date.isoformat() + 'Z',
                      end_date.isoformat() + 'Z', search_str='[scheduler] ')

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(parents=[tools.argparser])
    subparsers = parser.add_subparsers()

    parser_gen = subparsers.add_parser('generate', help='generate a schedule')
    parser_gen.add_argument('config')
    parser_gen.add_argument('year', type=int)
    parser_gen.add_argument('--calendar', type=str, default=None)
    parser_gen.add_argument('--publish', action='store_true', default=False)
    parser_gen.add_argument('--blocks', type=int)
    parser_gen.set_defaults(func=generate_schedule)

    parser_clear = subparsers.add_parser('clear', help='clear generated schedule')
    parser_clear.add_argument('calendar', type=str)
    parser_clear.add_argument('year', type=int)
    parser_clear.set_defaults(func=clear_schedule)

    args = parser.parse_args()

    start_time = time.clock()
    args.func(args)
    print('time = {} seconds'.format(time.clock() - start_time))
