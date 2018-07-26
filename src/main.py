import os
import time
from datetime import datetime, timedelta

from openpyxl import Workbook

import scheduler
from constants import *
from netflow import FlowNetwork
from settings import SettingsManager
from strategy import ConsecutiveStrategy, RandomStrategy


def publish_sched(sched, api, year):
    """
    Publishes the block and weekend assignments as events to Google 
    calendar.
    """
    for division in sched.divisions.values():
        for block_num in range(len(division.assignments)):
            clinician = division.assignments[block_num]
            for j in range(BLOCK_SIZE, 0, -1):
                week_start = datetime.strptime(
                    '{0}/{1:02d}/1/08:00/'.format(
                        year,
                        BLOCK_SIZE * (block_num + 1) - (j - 1)),
                    '%G/%V/%u/%H:%M/')
                week_end = week_start + timedelta(hours=WEEK_HOURS)
                summary = '{} - DIV:{} on call'.format(
                    clinician.name, division.name)
                api.create_event(
                    week_start.isoformat(),
                    week_end.isoformat(),
                    [clinician.email],
                    summary
                )

    for clinician in sched.clinicians.values():
        for week_num in clinician.weekends_assigned:
            weekend_start = datetime.strptime(
                '{0}/{1:02d}/5/17:00/'.format(
                    year,
                    week_num),
                '%G/%V/%u/%H:%M/')
            weekend_end = weekend_start + timedelta(hours=WEEKEND_HOURS)
            summary = '{} - on call'.format(clinician.name)
            api.create_event(
                weekend_start.isoformat(),
                weekend_end.isoformat(),
                [clinician.email],
                summary
            )


def write_to_excel(scheduler, year):
    wb = Workbook()
    sheet = wb.active
    col_index = 0
    divisions = list(scheduler.divisions.values())

    for col_index in range(len(divisions)):
        division = divisions[col_index]
        sheet.cell(1, col_index + 1, division.name)
        for row_index in range(len(division.assignments)):
            clinician = division.assignments[row_index]
            sheet.cell(2 * (row_index + 1), col_index + 1, clinician.name)
            sheet.cell(2 * (row_index + 1) + 1, col_index + 1, clinician.name)

    weekend_assignments = [
        _ for clinician in scheduler.clinicians.values()
        for _ in clinician.get_weekend_vars(
            lambda x: x.get_value() == 1.0
        )
    ]
    weekend_assignments.sort(key=lambda x: x.week_num)
    col_index += 1
    sheet.cell(1, col_index + 1, "Weekends")
    for wa in weekend_assignments:
        if wa.week_num in scheduler.long_weekends:
            sheet.cell((wa.week_num + 1), col_index +
                       1, wa.clinician.name + "****")
        else:
            sheet.cell((wa.week_num + 1), col_index + 1, wa.clinician.name)

    wb.save('{}-schedule.xlsx'.format(year))


def main(config_path, year, calendar_id=None, publish=False):
    sched = scheduler.Scheduler(config_path)
    events = []
    long_weekends = []

    if calendar_id:
        print("Retrieving {} calendar events from {}...".format(year, calendar_id))
        api = scheduler.API(calendar_id)
        start_date = datetime(year, 1, 1)
        end_date = start_date + timedelta(weeks=52)
        events = api.get_events(
            start_date.isoformat() + 'Z', end_date.isoformat() + 'Z')
        # TODO: read long weekends from api

    print("Populating scheduler...")
    sched.set_timeoff(events)
    sched.set_long_weekends(long_weekends)
    sched.build_lp()
    if sched.solve_lp():
        print("Found a feasible schedule!")
        sched.assign_schedule()
        write_to_excel(sched, year)

        if publish:
            print("Publishing {} schedule to {}...".format(year, calendar_id))
            publish_sched(sched, api, year)
    else:
        print("ERROR: Could not find a feasible schedule.")
        print("Try adjusting min/max values for clinicians.")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('config')
    parser.add_argument('year', type=int)
    parser.add_argument('--calendar', type=str, default=None)
    parser.add_argument('--publish', action='store_true', default=False)
    args = parser.parse_args()

    start_time = time.clock()
    main(args.config, args.year, args.calendar, args.publish)
    print('time = {} seconds'.format(time.clock() - start_time))
